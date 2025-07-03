
from .models import Job, Breakdown, Borrow
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .forms import JobForm
from django.core.mail import send_mail
from .forms import BreakdownForm
from .forms import BorrowForm
from django.conf import settings
from django.shortcuts import get_object_or_404
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages  # Optional, for user messages
from django.utils.dateparse import parse_date
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.shortcuts import redirect
from django.http import HttpResponseForbidden
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from tracking.models import Job, Breakdown, Borrow
import re
from django.core.cache import cache
import openai
from thefuzz import fuzz, process
import spacy
from datetime import datetime, timedelta
from openai import OpenAI
import os
from azure.ai.inference import ChatCompletionsClient
from azure.ai.inference.models import SystemMessage, UserMessage
from azure.core.credentials import AzureKeyCredential
from tracking.models import Job  # import your model
from django.db.models import Count
from .models import BackupPlan  
from .forms import BackupPlanForm
from collections import defaultdict

@login_required
def post_login_redirect(request):
    user = request.user
    if user.is_staff or user.is_superuser:
        return redirect('/admin/')
    else:
        return redirect('dashboard')


def get_status_note(status):
    return {
        'pending': 'Note: This job is still pending.',
        'finished': 'The job has been completed successfully.',
        'rejected': 'This job was rejected and marked for disposal.',
        'purchase': 'Items have been ordered for this job.',
    }.get(status, '')
    
@login_required
def dashboard_view(request):
    now = timezone.now()
    return render(request, 'dashboard.html', {
        'user': request.user,
        'current_time': now,
    })

@login_required
def job_list(request):
    jobs = Job.objects.all()
    return render(request, 'job_list.html', {'jobs': jobs})

@login_required
def job_report(request):
    jobs = Job.objects.all()

    # Filter params, matching your form's input names
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    center = request.GET.get('center')
    job_number = request.GET.get('job_number')
    serial_number = request.GET.get('serial_number')
    status = request.GET.get('status')

    # Filtering logic
    if date_from:
        jobs = jobs.filter(job_date__gte=parse_date(date_from))
    if date_to:
        jobs = jobs.filter(job_date__lte=parse_date(date_to))
    if center:
        jobs = jobs.filter(center__icontains=center)
    if job_number:
        jobs = jobs.filter(job_number__icontains=job_number)
    if serial_number:
        jobs = jobs.filter(serial_number__icontains=serial_number)
    if status:
        jobs = jobs.filter(status=status)

    context = {
        'jobs': jobs,
        'current_user': request.user,
        'filters': {
            'date_from': date_from or '',
            'date_to': date_to or '',
            'center': center or '',
            'job_number': job_number or '',
            'serial_number': serial_number or '',
            'status': status or '',
        }
    }
    return render(request, 'job_report.html', context)

@login_required
def delete_job(request, job_id):
    job = get_object_or_404(Job, id=job_id)

    if request.user != job.created_by and not request.user.is_superuser:
        messages.error(request, "You do not have permission to delete this job.")
        return redirect('job_report')

    # Delete immediately without confirmation page
    job.delete()
    messages.success(request, 'Job deleted successfully.')
    return redirect('job_report')

@login_required
def delete_selected_jobs(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_jobs')
        if not selected_ids:
            messages.warning(request, "No jobs selected to delete.")
            return redirect('job_report')

        jobs_to_delete = Job.objects.filter(id__in=selected_ids)
        deletable_jobs = [job for job in jobs_to_delete if job.created_by == request.user or request.user.is_superuser]

        if not deletable_jobs:
            messages.error(request, "You do not have permission to delete selected jobs.")
            return redirect('job_report')

        count = len(deletable_jobs)
        for job in deletable_jobs:
            job.delete()

        messages.success(request, f"Successfully deleted {count} job(s).")
        return redirect('job_report')

    return redirect('job_report')

@login_required
def breakdown_list(request):
    breakdowns = Breakdown.objects.all()
    return render(request, 'breakdown_list.html', {'breakdowns': breakdowns})

@login_required
def breakdown_report(request):
    query = request.GET.get('q')
    breakdowns = Breakdown.objects.all()

    if query:
        breakdowns = breakdowns.filter(
            Q(center__icontains=query) |
            Q(issue__icontains=query) |
            Q(job_number__icontains=query)
        )

    return render(request, 'breakdown_report.html', {'breakdowns': breakdowns})

@login_required
def borrow_list(request):
    borrows = Borrow.objects.all()
    return render(request, 'borrow_list.html', {'borrows': borrows})

@login_required
def borrow_report(request):
    def safe_parse_date(date_str):
        if isinstance(date_str, str) and date_str.strip():
            return parse_date(date_str)
        return None

    date_from = safe_parse_date(request.GET.get('date_from'))
    date_to = safe_parse_date(request.GET.get('date_to'))
    name = request.GET.get('name')
    department = request.GET.get('department')
    item_type = request.GET.get('item_type')

    borrows = Borrow.objects.all()

    if date_from:
        borrows = borrows.filter(date__gte=date_from)
    if date_to:
        borrows = borrows.filter(date__lte=date_to)
    if name:
        borrows = borrows.filter(name__icontains=name)
    if department:
        borrows = borrows.filter(department__icontains=department)
    if item_type:
        borrows = borrows.filter(item_type__icontains=item_type)

    # Handle bulk delete
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_borrows')
        if selected_ids:
            Borrow.objects.filter(id__in=selected_ids).delete()
            return redirect('borrow_report')

    context = {
        'borrows': borrows,
        'request': request,
        'current_user': request.user,
    }
    return render(request, 'tracking/borrow_report.html', context)

@login_required
def edit_borrow(request, pk):
    borrow = get_object_or_404(Borrow, pk=pk)
    if request.method == 'POST':
        form = BorrowForm(request.POST, instance=borrow)
        if form.is_valid():
            form.save()
            return redirect('borrow_report')
    else:
        form = BorrowForm(instance=borrow)
    return render(request, 'tracking/edit_borrow.html', {'form': form})

@login_required
def delete_borrow(request, pk):
    borrow = get_object_or_404(Borrow, pk=pk)

    # Permission check: only allow if user is admin or the creator
    if not request.user.is_superuser and borrow.created_by != request.user:
        return HttpResponseForbidden("You do not have permission to delete this entry.")

    borrow.delete()
    return redirect('borrow_report')

@login_required
def export_borrow_excel(request):
    borrows = Borrow.objects.all()

    # Example filters - adjust if you have filtering in borrow_report view
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    name = request.GET.get('name')
    department = request.GET.get('department')
    item_type = request.GET.get('item_type')

    if date_from:
        borrows = borrows.filter(date__gte=date_from)
    if date_to:
        borrows = borrows.filter(date__lte=date_to)
    if name:
        borrows = borrows.filter(name__icontains=name)
    if department:
        borrows = borrows.filter(department__icontains=department)
    if item_type:
        borrows = borrows.filter(item_type__icontains=item_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Borrower Report"

    headers = ['Date', 'Name', 'Designation', 'Department', 'Item Type', 'Days', 'Email', 'Reason', 'HOD Email']

    # Styling definitions same as above
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border

    fill_gray = PatternFill("solid", fgColor="F2F2F2")
    date_columns = [1]  # 'Date' is 1st column

    for row_num, b in enumerate(borrows, 2):
        fill = fill_gray if row_num % 2 == 0 else None
        values = [
            b.date.strftime('%Y-%m-%d') if b.date else '',
            b.name,
            b.designation,
            b.department,
            b.item_type,
            b.days,
            b.email,
            b.reason,
            getattr(b, 'hod_email', ''),  # adjust field name if different
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if col_num in date_columns and value:
                cell.number_format = "YYYY-MM-DD"
            if fill:
                cell.fill = fill
            cell.alignment = alignment_center if col_num in date_columns else Alignment(horizontal="left", vertical="center")

    # Auto width adjustment
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5

    # Add Excel table style
    last_row = ws.max_row
    last_col = ws.max_column
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName="BorrowerReport", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=borrower_report.xlsx'

    wb.save(response)
    return response

@login_required
def add_breakdown(request):
    if request.method == 'POST':
        form = BreakdownForm(request.POST)
        if form.is_valid():
            breakdown = form.save(commit=False)
            breakdown.created_by = request.user
            breakdown.save()
            return redirect('breakdown_report')
    else:
        form = BreakdownForm()
    return render(request, 'tracking/add_breakdown.html', {'form': form})
    
@login_required
def add_borrow(request):
    if request.method == 'POST':
        form = BorrowForm(request.POST)
        if form.is_valid():
            borrow = form.save(commit=False)
            borrow.created_by = request.user
            borrow.save()

            # Email to borrower
            subject = f"[Asset Checkout] {borrow.item_type} Borrowed by {borrow.name}"
            message = render_to_string('emails/borrower_notification.html', {
                'borrow': borrow
            })

            if borrow.email:
                borrower_email = EmailMultiAlternatives(
                    subject,
                    body=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[borrow.email]
                )
                borrower_email.attach_alternative(message, "text/html")
                borrower_email.send()

            # If borrower is not HOD and HOD email exists
            if borrow.designation.lower() != 'hod' and hasattr(borrow, 'hod_email') and borrow.hod_email:
                hod_message = render_to_string('emails/hod_notification.html', {
                    'borrow': borrow
                })

                hod_email = EmailMultiAlternatives(
                    subject=f"[Department Borrow Alert] {borrow.department} borrowed {borrow.item_type}",
                    body=hod_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[borrow.hod_email]
                )
                hod_email.attach_alternative(hod_message, "text/html")
                hod_email.send()

            return redirect('borrow_report')
    else:
        form = BorrowForm()
    return render(request, 'tracking/add_borrow.html', {'form': form})

@login_required
def add_job(request):
    if request.method == 'POST':
        form = JobForm(request.POST)
        if form.is_valid():
            job = form.save(commit=False)
            job.created_by = request.user  # Set the creator
            job.save()

            # Prepare email content
            status = job.status.lower()
            subject = ''
            message = ''

            # Common job info to include in all emails
            job_info = (
                f"Job Date: {job.job_date}\n"
                f"Center: {job.center}\n"
                f"Job Number: {job.job_number}\n"
                f"Item Type: {job.item_type}\n"
                f"Request By: {job.request_by}\n"
                f"Job Assignee: {job.job_assignee}\n"
                f"Center Sent Date: {job.center_sent_date}\n"
                f"Head Office Receive Date: {job.head_office_receive_date}\n"
                f"Serial Number: {job.serial_number}\n"
                f"Remark: {job.remark if job.remark else '-'}\n"
            )

            # Status-specific content
            if status == 'pending':
                subject = f"Job {job.job_number} - Pending"
                message = f"{job_info}\nStatus: Pending\nNote: This job is still pending."
            elif status == 'finished':
                subject = f"Job {job.job_number} - Completed"
                message = f"{job_info}\nStatus: Finished\nThe job has been completed successfully."
            elif status == 'rejected':
                subject = f"Job {job.job_number} - Rejected for Disposal"
                message = f"{job_info}\nStatus: Rejected\nThis job was rejected and marked for disposal."
            elif status == 'purchase':
                subject = f"Job {job.job_number} - Items Ordered"
                message = f"{job_info}\nStatus: Purchase\nItems have been ordered for this job."

  # Render HTML email
            html_content = render_to_string('emails/job_status_email.html', {
                'job': job,
                'status': status.title(),
                'note': get_status_note(status),
            })

            # Send email
            if subject:
                email = EmailMultiAlternatives(
                    subject=subject,
                    body=html_content,  # fallback plain body (can be improved)
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[job.area_manager_email],
                )
                email.attach_alternative(html_content, "text/html")
                email.send()

            return redirect('job_report')
    else:
        form = JobForm()

    return render(request, 'tracking/add_job.html', {'form': form})


@login_required
def edit_job(request, job_id):
    job = get_object_or_404(Job, id=job_id)
    old_status = job.status.lower()

    if request.method == 'POST':
        form = JobForm(request.POST, instance=job)
        if form.is_valid():
            updated_job = form.save()
            new_status = updated_job.status.lower()

            if new_status != old_status:
                job_info = (
                    f"Job Date: {updated_job.job_date}\n"
                    f"Center: {updated_job.center}\n"
                    f"Job Number: {updated_job.job_number}\n"
                    f"Item Type: {updated_job.item_type}\n"
                    f"Request By: {updated_job.request_by}\n"
                    f"Job Assignee: {updated_job.job_assignee}\n"
                    f"Center Sent Date: {updated_job.center_sent_date}\n"
                    f"Head Office Receive Date: {updated_job.head_office_receive_date}\n"
                    f"Serial Number: {updated_job.serial_number}\n"
                    f"Remark: {updated_job.remark if updated_job.remark else '-'}\n"
                )

                subject = ''
                message = ''

                if new_status == 'pending':
                    subject = f"Job {updated_job.job_number} - Pending"
                    message = f"{job_info}\nStatus: Pending\nNote: This job is still pending."
                elif new_status == 'finished':
                    subject = f"Job {updated_job.job_number} - Completed"
                    message = f"{job_info}\nStatus: Finished\nThe job has been completed successfully."
                elif new_status == 'rejected':
                    subject = f"Job {updated_job.job_number} - Rejected for Disposal"
                    message = f"{job_info}\nStatus: Rejected\nThis job was rejected and marked for disposal."
                elif new_status == 'purchase':
                    subject = f"Job {updated_job.job_number} - Items Ordered"
                    message = f"{job_info}\nStatus: Purchase\nItems have been ordered for this job."

                if subject and message:
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        [updated_job.area_manager_email],
                        fail_silently=False,
                    )

            return redirect('job_report')
    else:
        form = JobForm(instance=job)
    return render(request, 'tracking/edit_job.html', {'form': form, 'job': job})




@login_required
def export_job_excel(request):
    # Apply same filters as in job_report view
    jobs = Job.objects.all()

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    center = request.GET.get('center')
    job_number = request.GET.get('job_number')
    serial_number = request.GET.get('serial_number')
    status = request.GET.get('status')

    if date_from:
        jobs = jobs.filter(job_date__gte=date_from)
    if date_to:
        jobs = jobs.filter(job_date__lte=date_to)
    if center:
        jobs = jobs.filter(center__icontains=center)
    if job_number:
        jobs = jobs.filter(job_number__icontains=job_number)
    if serial_number:
        jobs = jobs.filter(serial_number__icontains=serial_number)
    if status:
        jobs = jobs.filter(status=status)

    # Create Workbook and Sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Report"

    # Define headers matching your table columns
    headers = [
        "Job Date", "Center", "Area Manager", "Area Manager Email", "Job Number",
        "Item Type", "Request By", "Requester Designation", "Job Assignee","Pronto No Receive",
        "Center Sent Date", "Head Office Receive Date", "Serial Number",
        "Finish Date", "Status", "Remark", "Created By"
    ]

    # Write header row with style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # nice blue
    alignment_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border

    # Write data rows with alternate fill colors
    fill_gray = PatternFill("solid", fgColor="F2F2F2")

    date_columns = [1, 10, 11, 15, 16, 17]  # columns with date fields

    for row_num, job in enumerate(jobs, 2):
        fill = fill_gray if row_num % 2 == 0 else None
        values = [
            job.job_date,
            job.center,
            job.area_manager,
            job.area_manager_email,
            job.job_number,
            job.item_type,
            job.request_by,
            job.requester_designation,
            job.job_assignee,
            job.pronto_no_receive,
            job.center_sent_date,
            job.head_office_receive_date,
            job.serial_number,
            job.finish_date if job.finish_date else "-",
            job.status,
            job.remark if job.remark else "-",
            job.created_by.username if job.created_by else "-"
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            # Format date columns
            if col_num in date_columns and value != "-":
                cell.number_format = "YYYY-MM-DD"
            # Apply fill for alternate rows
            if fill:
                cell.fill = fill
            # Align text left except dates centered
            if col_num in date_columns:
                cell.alignment = alignment_center
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths based on max length in column
    for col_num, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
        adjusted_width = max_length + 5
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add table style (Excel "Table" object for better UI)
    last_row = ws.max_row
    last_col = ws.max_column
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"

    tab = Table(displayName="JobReportTable", ref=table_ref)

    style = TableStyleInfo(
        name="TableStyleMedium9",  # Blue medium style
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Prepare response for download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=job_report.xlsx'

    wb.save(response)
    return response

@login_required
def breakdown_report(request):
    breakdowns = Breakdown.objects.all()

    # Filters example (adjust fields as per your model)
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    center = request.GET.get('center')
    job_number = request.GET.get('job_number')  # example field
    created_by = request.GET.get('created_by')  # example field

    if date_from:
        breakdowns = breakdowns.filter(date__gte=date_from)
    if date_to:
        breakdowns = breakdowns.filter(date__lte=date_to)
    if center:
        breakdowns = breakdowns.filter(center__icontains=center)
    if job_number:
        breakdowns = breakdowns.filter(job_number__icontains=job_number)
    if created_by:
        breakdowns = breakdowns.filter(created_by__username__icontains=created_by)

    context = {
        'breakdowns': breakdowns,
        'current_user': request.user,
        'filters': {
            'date_from': date_from or '',
            'date_to': date_to or '',
            'center': center or '',
            'job_number': job_number or '',
            'created_by': created_by or '',
        }
    }
    return render(request, 'tracking/breakdown_report.html', context)


@login_required
def delete_breakdown(request, breakdown_id):
    breakdown = get_object_or_404(Breakdown, id=breakdown_id)

    if request.user != breakdown.created_by and not request.user.is_superuser:
        messages.error(request, "You do not have permission to delete this breakdown.")
        return redirect('breakdown_report')

    breakdown.delete()
    messages.success(request, 'Breakdown deleted successfully.')
    return redirect('breakdown_report')


@login_required
def delete_selected_breakdowns(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_breakdowns')
        if not selected_ids:
            messages.warning(request, "No breakdowns selected to delete.")
            return redirect('breakdown_report')

        breakdowns_to_delete = Breakdown.objects.filter(id__in=selected_ids)
        deletable_breakdowns = [bd for bd in breakdowns_to_delete if bd.created_by == request.user or request.user.is_superuser]

        if not deletable_breakdowns:
            messages.error(request, "You do not have permission to delete selected breakdowns.")
            return redirect('breakdown_report')

        count = len(deletable_breakdowns)
        for bd in deletable_breakdowns:
            bd.delete()

        messages.success(request, f"Successfully deleted {count} breakdown(s).")
        return redirect('breakdown_report')

    return redirect('breakdown_report')


@login_required
def export_breakdown_excel(request):
    breakdowns = Breakdown.objects.all()

    # Filters - adjust field names if different
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    center = request.GET.get('center')
    job_number = request.GET.get('job_number')

    if start_date:
        breakdowns = breakdowns.filter(date__gte=start_date)
    if end_date:
        breakdowns = breakdowns.filter(date__lte=end_date)
    if center:
        breakdowns = breakdowns.filter(center__icontains=center)
    if job_number:
        breakdowns = breakdowns.filter(job_number__icontains=job_number)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Breakdown Report"

    headers = ['Date', 'Center', 'Job Number', 'Issue', 'Job Assignee', 'Comment', 'Created By']

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border

    fill_gray = PatternFill("solid", fgColor="F2F2F2")
    date_columns = [1]  # 'Date' is first column

    for row_num, b in enumerate(breakdowns, 2):
        fill = fill_gray if row_num % 2 == 0 else None
        values = [
            b.date.strftime('%Y-%m-%d') if b.date else '',
            b.center,
            b.job_number if hasattr(b, 'job_number') else '',
            b.issue,
            b.job_assignee,
            b.comment,
            b.created_by.username if b.created_by else '',
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if col_num in date_columns and value:
                cell.number_format = "YYYY-MM-DD"
            if fill:
                cell.fill = fill
            cell.alignment = alignment_center if col_num in date_columns else Alignment(horizontal="left", vertical="center")

    # Auto column widths
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 5

    # Add table style
    last_row = ws.max_row
    last_col = ws.max_column
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName="BreakdownReport", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=breakdown_report.xlsx'

    wb.save(response)
    return response


@login_required
def edit_breakdown(request, breakdown_id):
    breakdown = get_object_or_404(Breakdown, id=breakdown_id)
    # Save old status if you have a status field and want to handle changes/notifications like in job

    if request.method == 'POST':
        form = BreakdownForm(request.POST, instance=breakdown)
        if form.is_valid():
            updated_breakdown = form.save()
            # You can add status change notification logic here if you want

            return redirect('breakdown_report')
    else:
        form = BreakdownForm(instance=breakdown)

    return render(request, 'tracking/edit_breakdown.html', {'form': form, 'breakdown': breakdown})

nlp = spacy.load("en_core_web_sm")

# Known commands / intents for fuzzy matching
COMMANDS = [
    "show job",
    "pending jobs",
    "show breakdowns",
    "borrower details",
    "help",
    "summary for center",
]

def fuzzy_match_command(user_message):
    best_match, score = process.extractOne(user_message, COMMANDS, scorer=fuzz.token_sort_ratio)
    if score > 70:
        return best_match
    return None

def parse_user_message(message):
    doc = nlp(message.lower())
    intents = []
    entities = {}

    # Basic intent detection by keywords
    if "job" in message:
        intents.append("show_job")
    if "pending" in message:
        intents.append("pending_jobs")
    if "breakdown" in message:
        intents.append("show_breakdowns")
    if "borrow" in message:
        intents.append("borrower_details")
    if "help" in message or message.strip() == "/help":
        intents.append("help")
    if "summary" in message and "center" in message:
        intents.append("summary_for_center")

    # Extract entities - job numbers, centers, dates etc.
    for ent in doc.ents:
        entities[ent.label_] = ent.text

    # Also extract job number pattern like ECO-1021, eco1021, etc.
    import re
    job_match = re.search(r"eco-?\d+", message, re.IGNORECASE)
    if job_match:
        entities['JOB_NUMBER'] = job_match.group(0).upper().replace("-", "")

    # Extract center name heuristic (simple noun chunk or proper noun)
    for chunk in doc.noun_chunks:
        # You might want to refine this logic
        if "center" not in chunk.text and len(chunk.text) <= 20:
            entities['CENTER'] = chunk.text.title()

    return intents, entities

def format_job_details(job):
    # Format job details nicely with emojis and line breaks
    return (
        f"📋 *Job Details for {job.job_number}:*\n"
        f"- Job Date: {job.job_date.strftime('%b %d, %Y')}\n"
        f"- Center: {job.center}\n"
        f"- Area Manager: {job.area_manager}\n"
        f"- Area Manager Email: {job.area_manager_email}\n"
        f"- Item Type: {job.item_type}\n"
        f"- Request By: {job.request_by}\n"
        f"- Requester Designation: {job.requester_designation}\n"
        f"- Job Assignee: {job.job_assignee}\n"
        f"- Center Sent Date: {job.center_sent_date or 'N/A'}\n"
        f"- Head Office Receive Date: {job.head_office_receive_date or 'N/A'}\n"
        f"- Serial Number: {job.serial_number}\n"
        f"- Pronto No Receive: {job.pronto_no_receive or 'N/A'}\n"
        f"- Pronto No Sent: {job.pronto_no_sent or 'N/A'}\n"
        f"- Head Office Sent Date: {job.head_office_sent_date or 'N/A'}\n"
        f"- Center Receive Date: {job.center_receive_date or 'N/A'}\n"
        f"- Finish Date: {job.finish_date or 'N/A'}\n"
        f"- Status: {job.status}\n"
        f"- Remark: {job.remark or 'N/A'}"
    )

def format_breakdowns(breakdowns):
    if not breakdowns:
        return "⚠️ No breakdowns found."
    text = "🛠️ *Breakdown Reports:*\n"
    for b in breakdowns[:5]:  # limit to 5 latest
        text += (
            f"- Date: {b.date.strftime('%b %d, %Y')}, Center: {b.center}, Job#: {b.job_number}\n"
            f"  Issue: {b.issue}\n"
            f"  Assignee: {b.job_assignee}, Comment: {b.comment or 'N/A'}\n"
        )
    if len(breakdowns) > 5:
        text += f"...and {len(breakdowns)-5} more."
    return text

def format_borrowers(borrowers):
    if not borrowers:
        return "⚠️ No borrower records found."
    text = "📦 *Borrower Reports:*\n"
    for br in borrowers[:5]:  # limit to 5 latest
        text += (
            f"- Date: {br.date.strftime('%b %d, %Y')}, Name: {br.name}, Designation: {br.designation}\n"
            f"  Department: {br.department}, Item: {br.item_type}\n"
            f"  Days: {br.days}, Reason: {br.reason[:50]}...\n"
        )
    if len(borrowers) > 5:
        text += f"...and {len(borrowers)-5} more."
    return text

def get_pending_jobs_summary():
    cached = cache.get('pending_jobs_summary')
    if cached:
        return cached
    pending_jobs = Job.objects.filter(status__iexact="Pending")
    count = pending_jobs.count()
    summary = f"📊 Currently there are *{count}* pending jobs."
    cache.set('pending_jobs_summary', summary, timeout=300)  # cache 5 mins
    return summary

def get_summary_for_center(center):
    jobs = Job.objects.filter(center__iexact=center)
    if not jobs.exists():
        return f"⚠️ No jobs found for center '{center}'."
    pending = jobs.filter(status__iexact="Pending").count()
    finished = jobs.filter(status__iexact="Finished").count()
    rejected = jobs.filter(status__iexact="Rejected").count()
    purchase = jobs.filter(status__iexact="Purchase").count()
    return (
        f"📍 *Summary for Center {center}:*\n"
        f"- Pending: {pending}\n"
        f"- Finished: {finished}\n"
        f"- Rejected: {rejected}\n"
        f"- In Purchase: {purchase}"
    )



# Setup Azure inference client
endpoint = "https://models.inference.ai.azure.com"
model = "gpt-4o"
token = "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"

client = ChatCompletionsClient(
    endpoint=endpoint,
    credential=AzureKeyCredential(token),
)

@csrf_exempt
def chatbot_view(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)

    data = json.loads(request.body)
    user_message = data.get("message", "").lower()
    reply = "🤖 Sorry, I didn't understand that."

    try:
        job_match = re.search(r"job\s+([a-z0-9\-]+)", user_message)
        breakdown_match = re.search(r"breakdown.*?(\w+)", user_message)
        borrow_match = re.search(r"borrow(ed|er)?\s*(\w+)?", user_message)
        pending_match = re.search(r"pending\s+jobs", user_message)
        finished_match = re.search(r"finished\s+jobs", user_message)
        total_match = re.search(r"(total|how many|count).*(jobs?)", user_message)
        all_jobs_match = re.search(r"(show|list).*(all jobs?)", user_message)
        summary_match = re.search(r"summary.*center\s+(\w+)", user_message)
        help_match = re.match(r"/help", user_message)

        if help_match:
            reply = (
                "📚 **Available Commands:**\n\n"
                "• `/help` - Show this help message.\n"
                "• `job <job_number>` - Get details about a specific job.\n"
                "• `pending jobs` - List pending jobs.\n"
                "• `finished jobs` - List finished jobs.\n"
                "• `total jobs` or `how many jobs` - Show total job count.\n"
                "• `show all jobs` - List all jobs (limited to 20).\n"
                "• `breakdown <job_number>` - Show breakdown report for a job.\n"
                "• `borrow <item>` - Show borrow reports for an item.\n"
                "• `summary center <center_name>` - Show job summary for a center.\n"
            )

        elif job_match:
            job_no = job_match.group(1)
            job = Job.objects.filter(job_number__iexact=job_no).first()
            if job:
                reply = (
                    f"📋 **Job Details - {job.job_number}**\n"
                    f"────────────────────────────\n"
                    f"🗓 Date: {job.job_date}\n"
                    f"🏢 Center: {job.center}\n"
                    f"👤 Area Manager: {job.area_manager} ({job.area_manager_email})\n"
                    f"👥 Requested By: {job.request_by} ({job.requester_designation})\n"
                    f"🧑‍🔧 Assignee: {job.job_assignee}\n\n"
                    f"📦 Item Info:\n"
                    f"  • Serial No.: {job.serial_number}\n"
                    f"  • Pronto In: {job.pronto_no_receive}\n"
                    f"  • Pronto Out: {job.pronto_no_sent}\n\n"
                    f"📅 Timeline:\n"
                    f"  • Center Sent: {job.center_sent_date}\n"
                    f"  • HO Received: {job.head_office_receive_date}\n"
                    f"  • HO Sent: {job.head_office_sent_date}\n"
                    f"  • Center Received: {job.center_receive_date}\n"
                    f"  • Finish Date: {job.finish_date}\n\n"
                    f"📊 Status: {job.status}\n"
                    f"📝 Remark: {job.remark}\n"
                    f"👤 Created By: {job.created_by}\n"
                    "────────────────────────────"
                )
            else:
                reply = f"❌ No job found for: {job_no}"

        elif pending_match:
            jobs = Job.objects.filter(status__iexact="Pending")[:20]
            if jobs.exists():
                reply = "⏳ **Pending Jobs (up to 20 shown):**\n────────────────────────────\n"
                for job in jobs:
                    reply += f"• {job.job_number} | {job.center} | {job.item_type}\n"
                reply += "────────────────────────────"
            else:
                reply = "✅ No pending jobs found!"

        elif finished_match:
            jobs = Job.objects.filter(status__iexact="Finished")[:20]
            if jobs.exists():
                reply = "✅ **Finished Jobs (up to 20 shown):**\n────────────────────────────\n"
                for job in jobs:
                    reply += f"• {job.job_number} | {job.center} | {job.item_type}\n"
                reply += "────────────────────────────"
            else:
                reply = "❌ No finished jobs found."

        elif total_match:
            total_jobs = Job.objects.count()
            reply = f"📊 Total number of jobs in the system: **{total_jobs}**"

        elif all_jobs_match:
            jobs = Job.objects.all()[:20]
            if jobs.exists():
                reply = "📋 **All Jobs (up to 20 shown):**\n────────────────────────────\n"
                for job in jobs:
                    reply += f"• {job.job_number} | {job.center} | Status: {job.status}\n"
                reply += "────────────────────────────"
            else:
                reply = "❌ No jobs found."

        elif breakdown_match:
            job_no = breakdown_match.group(1)
            breakdowns = Breakdown.objects.filter(job_number__icontains=job_no)
            if breakdowns.exists():
                reply = f"🔧 **Breakdown Report for Job {job_no}:**\n────────────────────────────\n"
                for b in breakdowns:
                    reply += (
                        f"• Date: {b.date} | Center: {b.center}\n"
                        f"  Issue: {b.issue}\n"
                        f"  Assignee: {b.job_assignee}\n"
                        f"  Comment: {b.comment}\n"
                        "────────────────────────────\n"
                    )
            else:
                reply = f"❌ No breakdown reports found for job: {job_no}"

        elif borrow_match:
            item = borrow_match.group(2) or ""
            borrows = Borrow.objects.filter(item_type__icontains=item)
            if borrows.exists():
                reply = f"📦 **Borrow Reports for '{item}':**\n────────────────────────────\n"
                for b in borrows:
                    reply += (
                        f"• Date: {b.date}\n"
                        f"  Name: {b.name} | Designation: {b.designation} | Department: {b.department}\n"
                        f"  Item: {b.item_type} for {b.days} days\n"
                        f"  Reason: {b.reason}\n"
                        f"  Email: {b.email}\n"
                        f"  Created By: {b.created_by}\n"
                        "────────────────────────────\n"
                    )
            else:
                reply = f"❌ No borrow reports found related to: {item}"

        elif summary_match:
            center = summary_match.group(1)
            jobs = Job.objects.filter(center__icontains=center)
            total = jobs.count()
            pending = jobs.filter(status__iexact="Pending").count()
            finished = jobs.filter(status__iexact="Finished").count()
            reply = (
                f"📊 **Summary for Center: {center}**\n"
                "────────────────────────────\n"
                f"• Total Jobs: {total}\n"
                f"• Pending Jobs: {pending}\n"
                f"• Finished Jobs: {finished}\n"
                "────────────────────────────"
            )

        else:
            # Fallback to AI if no pattern matched
            ai_response = client.complete(
                model=model,
                messages=[
                    SystemMessage(content="You are a helpful assistant for job tracking."),
                    UserMessage(content=user_message)
                ],
                temperature=0.7,
                top_p=1,
            )
            reply = ai_response.choices[0].message.content.strip()

    except Exception as e:
        reply = f"❌ Error: {str(e)}"

    return JsonResponse({"reply": reply})


@login_required
def add_backupplan(request):
    if request.method == 'POST':
        form = BackupPlanForm(request.POST)
        if form.is_valid():
            backup_plan = form.save(commit=False)
            backup_plan.created_by = request.user  # track who created it
            backup_plan.save()
            return redirect('backupplan_list')
    else:
        form = BackupPlanForm()
    return render(request, 'tracking/backupplan/add_backupplan.html', {'form': form})



from django.shortcuts import render
from collections import defaultdict
from .models import BackupPlan

def backupplan_list(request):
    all_plans = BackupPlan.objects.all().order_by('-date')

    grouped = defaultdict(list)
    for plan in all_plans:
        grouped[plan.area_manager].append(plan)

    return render(request, 'tracking/backupplan/backupplan_list.html', {
        'backupplans': dict(grouped)
    })

from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from .models import BackupPlan
from .forms import BackupPlanForm

@login_required
def delete_backupplan(request, plan_id):
    plan = get_object_or_404(BackupPlan, id=plan_id)

    if request.user != plan.created_by and not request.user.is_superuser:
        messages.error(request, "You do not have permission to delete this backup plan.")
        return redirect('backupplan_list')

    plan.delete()
    messages.success(request, 'Backup plan deleted successfully.')
    return redirect('backupplan_list')


@login_required
def delete_selected_backupplans(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_backupplans')
        if not selected_ids:
            messages.warning(request, "No backup plans selected to delete.")
            return redirect('backupplan_list')

        plans_to_delete = BackupPlan.objects.filter(id__in=selected_ids)
        deletable_plans = [plan for plan in plans_to_delete if plan.created_by == request.user or request.user.is_superuser]

        if not deletable_plans:
            messages.error(request, "You do not have permission to delete the selected backup plans.")
            return redirect('backupplan_list')

        count = len(deletable_plans)
        for plan in deletable_plans:
            plan.delete()

        messages.success(request, f"Successfully deleted {count} backup plan(s).")
        return redirect('backupplan_list')

    return redirect('backupplan_list')
    


@login_required
def edit_backupplan(request, plan_id):
    plan = get_object_or_404(BackupPlan, id=plan_id)

    if request.user != plan.created_by and not request.user.is_superuser:
        messages.error(request, "You do not have permission to edit this backup plan.")
        return redirect('backupplan_list')

    if request.method == 'POST':
        form = BackupPlanForm(request.POST, instance=plan)
        if form.is_valid():
            form.save()
            messages.success(request, 'Backup plan updated successfully.')
            return redirect('backupplan_list')
    else:
        form = BackupPlanForm(instance=plan)

    return render(request, 'tracking/edit_backupplan.html', {'form': form, 'plan': plan})

def export_backupplans_excel(request):
    all_plans = BackupPlan.objects.all().order_by('area_manager', '-date')

    grouped = defaultdict(list)
    for plan in all_plans:
        grouped[plan.area_manager].append(plan)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Backup Plans"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Center Name", "Date", "Test Lane 1", "Test Lane 2", "Test Lane 3",
        "Registration Lane", "Certificate Lane", "Backup PC", "Total PCs", "Total Monitors",
        "Mini PC", "HP", "Eswis", "Fingerprint Machines", "Backup Fingerprint Machines",
        "UPS", "Backup UPS", "Wingle", "Dongle", "Octopus", "Remark", "Created By", "Actions"
    ]

    row_num = 1
    for manager, plans in grouped.items():
        # Area Manager Title Row
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
        cell = ws.cell(row=row_num, column=1)
        cell.value = f"🧑‍💼 Area Manager: {manager}"
        cell.font = Font(bold=True, size=12)
        cell.alignment = center_align
        row_num += 1

        # Header Row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        row_num += 1

        # Data Rows
        for plan in plans:
            row = [
                plan.center_name,
                plan.date.strftime("%b %d, %Y"),
                plan.testing_lane_1_pc or "-",
                plan.testing_lane_2_pc or "-",
                plan.testing_lane_3_pc or "-",
                plan.registration_lane_pc or "-",
                plan.certificate_lane_pc or "-",
                plan.backup_pc or "-",
                plan.total_pc,
                plan.total_monitors,
                1 if plan.testing_lane_1_pc == "Mini PC" else 0,
                1 if plan.testing_lane_2_pc == "HP" else 0,
                1 if plan.testing_lane_3_pc == "Eswis" else 0,
                plan.fingerprint_machines,
                plan.backup_fingerprint_machines,
                plan.ups,
                plan.backup_ups,
                plan.wingles,
                plan.dongles,
                plan.octopuses,
                plan.remark or "-",
                plan.created_by.username if plan.created_by else "",
                "Edit/Delete" if request.user == plan.created_by or request.user.is_superuser else "No Access"
            ]

            for col_num, val in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = val
                cell.alignment = center_align
            row_num += 1

    # Auto-size columns
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 2

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=backup_plans.xlsx"
    wb.save(response)
    return response